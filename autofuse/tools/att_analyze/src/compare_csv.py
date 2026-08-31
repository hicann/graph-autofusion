#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# ----------------------------------------------------------------------------------------------------------
import csv
import os
import argparse
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class ComparisonResult:
    file1_path: str
    file2_path: str
    file1_rows: int
    file2_rows: int
    file1_cols: List[str]
    file2_cols: List[str]
    common_operators: List[str]
    only_in_file1: List[str]
    only_in_file2: List[str]
    operator_differences: Dict[str, List["FieldDifference"]]
    structure_differences: List[str]
    performance_differences: List["PerformanceDifference"]


@dataclass
class FieldDifference:
    field_name: str
    value1: str
    value2: str
    diff_type: str
    numeric_diff: Optional[float] = None
    percentage_diff: Optional[float] = None


@dataclass
class PerformanceDifference:
    operator: str
    metric: str
    value1: float
    value2: float
    absolute_diff: float
    percentage_diff: float
    improvement: bool


class CSVComparator:
    @staticmethod
    def _row_key(row: Dict) -> str:
        """Identify the same graph/result/group across runs.

        ``Case`` is deliberately excluded: selecting a different tiling case is
        the difference this comparator is intended to show.
        """
        return "|".join(
            str(row.get(field, ""))
            for field in ("Operator", "Graph", "Result", "Group")
        )

    def __init__(self):
        self.performance_fields = [
            "AIV_MTE2",
            "AIV_MTE3",
            "Objective Value",
            "Result Perf",
        ]
        self.numeric_fields = [
            "Graph",
            "Result",
            "Group",
            "Case",
            "AIV_MTE2",
            "AIV_MTE3",
            "Objective Value",
            "Result Perf",
        ]

    def read_csv_file(self, file_path: str) -> Tuple[List[Dict[str, str]], List[str]]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"CSV file '{file_path}' not found.")

        rows = []
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                rows.append(row)

        return rows, fieldnames

    def compare_structure(
        self, file1_cols: List[str], file2_cols: List[str]
    ) -> List[str]:
        differences = []

        if file1_cols != file2_cols:
            only_in_file1 = set(file1_cols) - set(file2_cols)
            only_in_file2 = set(file2_cols) - set(file1_cols)

            if only_in_file1:
                differences.append(f"列仅在文件1中: {', '.join(only_in_file1)}")
            if only_in_file2:
                differences.append(f"列仅在文件2中: {', '.join(only_in_file2)}")

        return differences

    def compare_operators(
        self, file1_rows: List[Dict], file2_rows: List[Dict]
    ) -> Tuple[List[str], List[str], List[str]]:
        file1_operators = {self._row_key(row): row for row in file1_rows}
        file2_operators = {self._row_key(row): row for row in file2_rows}

        common_operators = set(file1_operators.keys()) & set(file2_operators.keys())
        only_in_file1 = set(file1_operators.keys()) - set(file2_operators.keys())
        only_in_file2 = set(file2_operators.keys()) - set(file1_operators.keys())

        return sorted(common_operators), sorted(only_in_file1), sorted(only_in_file2)

    def parse_numeric(self, value: str) -> Optional[float]:
        if value == "" or value == "N/A":
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def compare_field_values(
        self, value1: str, value2: str, field_name: str
    ) -> Optional[FieldDifference]:
        if value1 == value2:
            return None

        diff_type = "字符串差异"
        numeric_diff = None
        percentage_diff = None

        num1 = self.parse_numeric(value1)
        num2 = self.parse_numeric(value2)

        if num1 is not None and num2 is not None:
            diff_type = "数值差异"
            numeric_diff = num2 - num1
            if num1 != 0:
                percentage_diff = (numeric_diff / num1) * 100
            else:
                percentage_diff = 0.0 if num2 == 0 else float("inf")

        return FieldDifference(
            field_name=field_name,
            value1=value1,
            value2=value2,
            diff_type=diff_type,
            numeric_diff=numeric_diff,
            percentage_diff=percentage_diff,
        )

    def compare_operator_rows(
        self, row1: Dict, row2: Dict, operator: str
    ) -> List[FieldDifference]:
        differences = []

        all_fields = set(row1.keys()) | set(row2.keys())

        for field in all_fields:
            value1 = row1.get(field, "N/A")
            value2 = row2.get(field, "N/A")

            diff = self.compare_field_values(value1, value2, field)
            if diff:
                differences.append(diff)

        return differences

    def compare_performance_metrics(
        self,
        common_operators: List[str],
        file1_rows: List[Dict],
        file2_rows: List[Dict],
    ) -> List[PerformanceDifference]:
        file1_operators = {self._row_key(row): row for row in file1_rows}
        file2_operators = {self._row_key(row): row for row in file2_rows}

        performance_diffs = []

        for operator in common_operators:
            row1 = file1_operators[operator]
            row2 = file2_operators[operator]

            for metric in self.performance_fields:
                value1 = self.parse_numeric(row1.get(metric, ""))
                value2 = self.parse_numeric(row2.get(metric, ""))

                if value1 is not None and value2 is not None and value1 != value2:
                    absolute_diff = value2 - value1
                    percentage_diff = (
                        (absolute_diff / value1) * 100 if value1 != 0 else 0.0
                    )
                    improvement = absolute_diff < 0

                    performance_diffs.append(
                        PerformanceDifference(
                            operator=operator,
                            metric=metric,
                            value1=value1,
                            value2=value2,
                            absolute_diff=absolute_diff,
                            percentage_diff=percentage_diff,
                            improvement=improvement,
                        )
                    )

        return performance_diffs

    def compare_csv_files(self, file1_path: str, file2_path: str) -> ComparisonResult:
        file1_rows, file1_cols = self.read_csv_file(file1_path)
        file2_rows, file2_cols = self.read_csv_file(file2_path)

        structure_differences = self.compare_structure(file1_cols, file2_cols)

        common_operators, only_in_file1, only_in_file2 = self.compare_operators(
            file1_rows, file2_rows
        )

        file1_operators = {self._row_key(row): row for row in file1_rows}
        file2_operators = {self._row_key(row): row for row in file2_rows}

        operator_differences = {}
        for operator in common_operators:
            row1 = file1_operators[operator]
            row2 = file2_operators[operator]
            diffs = self.compare_operator_rows(row1, row2, operator)
            if diffs:
                operator_differences[operator] = diffs

        performance_differences = self.compare_performance_metrics(
            common_operators, file1_rows, file2_rows
        )

        return ComparisonResult(
            file1_path=file1_path,
            file2_path=file2_path,
            file1_rows=len(file1_rows),
            file2_rows=len(file2_rows),
            file1_cols=file1_cols,
            file2_cols=file2_cols,
            common_operators=common_operators,
            only_in_file1=only_in_file1,
            only_in_file2=only_in_file2,
            operator_differences=operator_differences,
            structure_differences=structure_differences,
            performance_differences=performance_differences,
        )


class ComparisonReporter:
    def __init__(self, result: ComparisonResult):
        self.result = result

    def print_console_report(self):
        self._print_header()
        self._print_file_info()
        self._print_structure_comparison()
        self._print_operator_comparison()
        self._print_detailed_differences()
        self._print_performance_comparison()
        self._print_summary()

    def _print_header(self):
        print("=" * 80)
        print("CSV文件对比分析报告")
        print("=" * 80)
        print()

    def _print_file_info(self):
        print("文件信息：")
        print(f"  文件1: {self.result.file1_path} ({self.result.file1_rows} 个算子)")
        print(f"  文件2: {self.result.file2_path} ({self.result.file2_rows} 个算子)")
        print()

    def _print_structure_comparison(self):
        print("结构对比：")
        print(
            f"  列数: {len(self.result.file1_cols)} vs {len(self.result.file2_cols)}",
            end="",
        )
        col_diff = len(self.result.file2_cols) - len(self.result.file1_cols)
        if col_diff != 0:
            print(f" (差异: {col_diff:+d})")
        else:
            print(" (相同)")

        print(f"  行数: {self.result.file1_rows} vs {self.result.file2_rows}", end="")
        row_diff = self.result.file2_rows - self.result.file1_rows
        if row_diff != 0:
            print(f" (差异: {row_diff:+d})")
        else:
            print(" (相同)")

        if self.result.structure_differences:
            print("\n  结构差异：")
            for diff in self.result.structure_differences:
                print(f"    - {diff}")

        print()

    def _print_operator_comparison(self):
        print("算子对比：")
        print(f"  共同算子: {len(self.result.common_operators)} 个")
        if self.result.common_operators:
            print(f"    {', '.join(self.result.common_operators)}")

        if self.result.only_in_file1:
            print(f"  仅在文件1中: {len(self.result.only_in_file1)} 个")
            print(f"    {', '.join(self.result.only_in_file1)}")

        if self.result.only_in_file2:
            print(f"  仅在文件2中: {len(self.result.only_in_file2)} 个")
            print(f"    {', '.join(self.result.only_in_file2)}")

        print()

    def _print_detailed_differences(self):
        if not self.result.operator_differences:
            print("详细差异分析：")
            print("  无差异")
            print()
            return

        print("详细差异分析：")
        for operator, diffs in self.result.operator_differences.items():
            print(f"  [{operator}]")
            for diff in diffs:
                if diff.numeric_diff is not None:
                    print(
                        f"    - {diff.field_name}: {diff.value1} → {diff.value2}",
                        end="",
                    )
                    if diff.percentage_diff is not None and abs(
                        diff.percentage_diff
                    ) != float("inf"):
                        print(
                            f" (差异: {diff.numeric_diff:+.6f}, {diff.percentage_diff:+.2f}%)"
                        )
                    else:
                        print(f" (差异: {diff.numeric_diff:+.6f})")
                else:
                    print(f"    - {diff.field_name}: {diff.value1} → {diff.value2}")
        print()

    def _print_performance_comparison(self):
        if not self.result.performance_differences:
            return

        print("性能指标对比：")
        for perf_diff in self.result.performance_differences:
            status = "✓ 改善" if perf_diff.improvement else "✗ 下降"
            print(f"  [{perf_diff.operator}] {perf_diff.metric}")
            print(f"    {perf_diff.value1:.6f} → {perf_diff.value2:.6f}", end="")
            print(
                f" ({perf_diff.absolute_diff:+.6f}, {perf_diff.percentage_diff:+.2f}%) {status}"
            )
        print()

    def _print_summary(self):
        print("=" * 80)
        print("对比总结：")
        total_differences = sum(
            len(diffs) for diffs in self.result.operator_differences.values()
        )
        print(f"  - 总计 {total_differences} 个字段差异")
        print(f"  - {len(self.result.performance_differences)} 个性能指标变化")

        if self.result.only_in_file1 or self.result.only_in_file2:
            print(
                f"  - {len(self.result.only_in_file1) + len(self.result.only_in_file2)} 个算子差异"
            )

        has_differences = (
            self.result.structure_differences
            or self.result.operator_differences
            or self.result.only_in_file1
            or self.result.only_in_file2
        )

        if not has_differences:
            print("  ✓ 两个文件完全相同")
        else:
            print("  ✓ 两个文件存在差异")

        print("=" * 80)

    def save_text_report(self, output_file: str):
        with open(output_file, "w", encoding="utf-8") as f:
            import sys

            original_stdout = sys.stdout
            sys.stdout = f

            self.print_console_report()

            sys.stdout = original_stdout

        print(f"对比报告已保存到: {output_file}")

    def save_excel_report(self, output_file: str):
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not installed. Falling back to text format.")
            text_file = output_file.replace(".xlsx", ".txt").replace(".xls", ".txt")
            self.save_text_report(text_file)
            return

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "摘要"
        self._write_summary_sheet(ws_summary)

        if self.result.operator_differences:
            ws_details = wb.create_sheet("详细差异")
            self._write_details_sheet(ws_details)

        if self.result.performance_differences:
            ws_performance = wb.create_sheet("性能对比")
            self._write_performance_sheet(ws_performance)

        wb.save(output_file)
        print(f"对比报告已保存到: {output_file}")

    def _write_summary_sheet(self, ws):
        headers = ["项目", "文件1", "文件2", "差异"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        ws.cell(row=row, column=1, value="文件路径")
        ws.cell(row=row, column=2, value=self.result.file1_path)
        ws.cell(row=row, column=3, value=self.result.file2_path)
        ws.cell(row=row, column=4, value="")

        row += 1
        ws.cell(row=row, column=1, value="算子数量")
        ws.cell(row=row, column=2, value=self.result.file1_rows)
        ws.cell(row=row, column=3, value=self.result.file2_rows)
        ws.cell(
            row=row,
            column=4,
            value=f"{self.result.file2_rows - self.result.file1_rows:+d}",
        )

        row += 1
        ws.cell(row=row, column=1, value="列数")
        ws.cell(row=row, column=2, value=len(self.result.file1_cols))
        ws.cell(row=row, column=3, value=len(self.result.file2_cols))
        ws.cell(
            row=row,
            column=4,
            value=f"{len(self.result.file2_cols) - len(self.result.file1_cols):+d}",
        )

        row += 1
        ws.cell(row=row, column=1, value="共同算子")
        ws.cell(row=row, column=2, value=len(self.result.common_operators))
        ws.cell(row=row, column=3, value=len(self.result.common_operators))
        ws.cell(row=row, column=4, value="相同")

        row += 1
        ws.cell(row=row, column=1, value="仅在文件1中")
        ws.cell(row=row, column=2, value=len(self.result.only_in_file1))
        ws.cell(row=row, column=3, value=0)
        ws.cell(row=row, column=4, value="")

        row += 1
        ws.cell(row=row, column=1, value="仅在文件2中")
        ws.cell(row=row, column=2, value=0)
        ws.cell(row=row, column=3, value=len(self.result.only_in_file2))
        ws.cell(row=row, column=4, value="")

        row += 1
        total_differences = sum(
            len(diffs) for diffs in self.result.operator_differences.values()
        )
        ws.cell(row=row, column=1, value="字段差异总数")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value=total_differences)

        row += 1
        ws.cell(row=row, column=1, value="性能指标变化数")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value=len(self.result.performance_differences))

        for col_num in range(1, 5):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    def _write_details_sheet(self, ws):
        headers = [
            "算子",
            "字段名",
            "文件1值",
            "文件2值",
            "差异类型",
            "数值差异",
            "百分比差异",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        for operator, diffs in self.result.operator_differences.items():
            for diff in diffs:
                ws.cell(row=row, column=1, value=operator)
                ws.cell(row=row, column=2, value=diff.field_name)
                ws.cell(row=row, column=3, value=diff.value1)
                ws.cell(row=row, column=4, value=diff.value2)
                ws.cell(row=row, column=5, value=diff.diff_type)

                if diff.numeric_diff is not None:
                    ws.cell(row=row, column=6, value=diff.numeric_diff)
                if diff.percentage_diff is not None and abs(
                    diff.percentage_diff
                ) != float("inf"):
                    ws.cell(row=row, column=7, value=diff.percentage_diff)

                row += 1

        for col_num in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

    def _write_performance_sheet(self, ws):
        headers = [
            "算子",
            "性能指标",
            "文件1值",
            "文件2值",
            "绝对差异",
            "百分比差异",
            "状态",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        for perf_diff in self.result.performance_differences:
            ws.cell(row=row, column=1, value=perf_diff.operator)
            ws.cell(row=row, column=2, value=perf_diff.metric)
            ws.cell(row=row, column=3, value=perf_diff.value1)
            ws.cell(row=row, column=4, value=perf_diff.value2)
            ws.cell(row=row, column=5, value=perf_diff.absolute_diff)
            ws.cell(row=row, column=6, value=perf_diff.percentage_diff)
            ws.cell(
                row=row, column=7, value="改善" if perf_diff.improvement else "下降"
            )
            row += 1

        for col_num in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15


def main():
    parser = argparse.ArgumentParser(
        description="Compare two CSV files generated by summary_templates.py"
    )
    parser.add_argument("file1", help="Path to first CSV file")
    parser.add_argument("file2", help="Path to second CSV file")
    parser.add_argument(
        "-o", "--output", help="Output file path (optional, default: print to console)"
    )
    parser.add_argument(
        "-f",
        "--format",
        choices=["console", "text", "excel"],
        default="console",
        help="Output format: console (default), text, or excel",
    )

    args = parser.parse_args()

    try:
        comparator = CSVComparator()
        result = comparator.compare_csv_files(args.file1, args.file2)

        reporter = ComparisonReporter(result)

        if args.format == "console":
            if args.output:
                reporter.save_text_report(args.output)
            else:
                reporter.print_console_report()
        elif args.format == "text":
            output_file = args.output if args.output else "comparison_report.txt"
            reporter.save_text_report(output_file)
        elif args.format == "excel":
            output_file = args.output if args.output else "comparison_report.xlsx"
            reporter.save_excel_report(output_file)

    except FileNotFoundError as e:
        print(f"错误: {e}")
    except Exception as e:
        print(f"错误: {e}")


if __name__ == "__main__":
    main()
