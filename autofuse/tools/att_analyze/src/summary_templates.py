#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# ----------------------------------------------------------------------------------------------------------
# 空壳转发：LogParser 和 OperatorSummary 已迁移到 core/log_parser.py
import os
import csv
import sys
from typing import List

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from core.log_parser import LogParser, OperatorSummary  # noqa: F401


def find_log_files(path):
    """递归查找目录下的所有日志文件"""
    log_files = []
    if os.path.isfile(path):
        return [path]
    elif os.path.isdir(path):
        for root, dirs, files in os.walk(path):
            dirs.sort()
            files.sort()
            for file in files:
                if file.endswith(".log"):
                    log_files.append(os.path.join(root, file))
        return sorted(log_files)
    else:
        return []


def print_summary_table(summaries: List[OperatorSummary]):
    if not summaries:
        print("No data to display.")
        return

    all_tiling_keys = set()
    for summary in summaries:
        all_tiling_keys.update(summary.tiling_values.keys())

    fixed_tiling_keys = ["ub_size", "block_dim"]
    performance_keys = ["AIV_MTE2", "AIV_MTE3"]
    dynamic_tiling_keys = sorted(
        all_tiling_keys - set(fixed_tiling_keys) - set(performance_keys)
    )
    all_columns = [
        "Operator",
        "Graph",
        "Result",
        "Group",
        "Case",
        "AIV_MTE2",
        "AIV_MTE3",
        "Objective Value",
        "Result Perf",
    ]
    all_columns.extend(dynamic_tiling_keys)
    all_columns.extend(fixed_tiling_keys)

    col_widths = {}
    for col in all_columns:
        col_widths[col] = max(len(col), 15)

    for summary in summaries:
        col_widths["Operator"] = max(col_widths["Operator"], len(summary.operator_name))

    header = "  ".join(col.ljust(col_widths[col]) for col in all_columns)
    print(header)
    print("-" * len(header))

    for summary in summaries:
        row = []
        row.append(summary.operator_name.ljust(col_widths["Operator"]))
        row.append(str(summary.graph).ljust(col_widths["Graph"]))
        row.append(str(summary.result).ljust(col_widths["Result"]))
        row.append(str(summary.group).ljust(col_widths["Group"]))
        row.append(str(summary.case).ljust(col_widths["Case"]))
        row.append(
            str(summary.aiv_mte2 if summary.aiv_mte2 is not None else "N/A").ljust(
                col_widths["AIV_MTE2"]
            )
        )
        row.append(
            str(summary.aiv_mte3 if summary.aiv_mte3 is not None else "N/A").ljust(
                col_widths["AIV_MTE3"]
            )
        )
        row.append(
            str(
                summary.objective_value
                if summary.objective_value is not None
                else "N/A"
            ).ljust(col_widths["Objective Value"])
        )
        row.append(
            str(
                summary.result_performance
                if summary.result_performance is not None
                else "N/A"
            ).ljust(col_widths["Result Perf"])
        )

        for key in dynamic_tiling_keys:
            value = summary.tiling_values.get(key, "N/A")
            row.append(str(value).ljust(col_widths[key]))

        for key in fixed_tiling_keys:
            value = summary.tiling_values.get(key, "N/A")
            row.append(str(value).ljust(col_widths[key]))

        print("  ".join(row))


def export_to_csv(summaries: List[OperatorSummary], output_file: str):
    if not summaries:
        print("No data to export.")
        return

    all_tiling_keys = set()
    for summary in summaries:
        all_tiling_keys.update(summary.tiling_values.keys())

    fixed_tiling_keys = ["ub_size", "block_dim"]
    performance_keys = ["AIV_MTE2", "AIV_MTE3"]
    dynamic_tiling_keys = sorted(
        all_tiling_keys - set(fixed_tiling_keys) - set(performance_keys)
    )

    fieldnames = [
        "Operator",
        "Graph",
        "Result",
        "Group",
        "Case",
        "AIV_MTE2",
        "AIV_MTE3",
        "Objective Value",
        "Result Perf",
    ]
    fieldnames.extend(dynamic_tiling_keys)
    fieldnames.extend(fixed_tiling_keys)

    with open(output_file, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()

        for summary in summaries:
            row = {
                "Operator": summary.operator_name,
                "Graph": summary.graph,
                "Result": summary.result,
                "Group": summary.group,
                "Case": summary.case,
                "AIV_MTE2": summary.aiv_mte2 if summary.aiv_mte2 is not None else "",
                "AIV_MTE3": summary.aiv_mte3 if summary.aiv_mte3 is not None else "",
                "Objective Value": summary.objective_value
                if summary.objective_value is not None
                else "",
                "Result Perf": summary.result_performance
                if summary.result_performance is not None
                else "",
            }

            for key in dynamic_tiling_keys:
                row[key] = summary.tiling_values.get(key, "")

            for key in fixed_tiling_keys:
                row[key] = summary.tiling_values.get(key, "")

            writer.writerow(row)


def export_to_excel(summaries: List[OperatorSummary], output_file: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Warning: openpyxl not installed. Falling back to CSV format.")
        csv_file = output_file.replace(".xlsx", ".csv").replace(".xls", ".csv")
        export_to_csv(summaries, csv_file)
        print(f"Data exported to {csv_file}")
        return

    if not summaries:
        print("No data to export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    all_tiling_keys = set()
    for summary in summaries:
        all_tiling_keys.update(summary.tiling_values.keys())

    fixed_tiling_keys = ["ub_size", "block_dim"]
    performance_keys = ["AIV_MTE2", "AIV_MTE3"]
    dynamic_tiling_keys = sorted(
        all_tiling_keys - set(fixed_tiling_keys) - set(performance_keys)
    )

    fieldnames = [
        "Operator",
        "Graph",
        "Result",
        "Group",
        "Case",
        "AIV_MTE2",
        "AIV_MTE3",
        "Objective Value",
        "Result Perf",
    ]
    fieldnames.extend(dynamic_tiling_keys)
    fieldnames.extend(fixed_tiling_keys)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, summary in enumerate(summaries, 2):
        data = {
            "Operator": summary.operator_name,
            "Graph": summary.graph,
            "Result": summary.result,
            "Group": summary.group,
            "Case": summary.case,
            "AIV_MTE2": summary.aiv_mte2 if summary.aiv_mte2 is not None else "",
            "AIV_MTE3": summary.aiv_mte3 if summary.aiv_mte3 is not None else "",
            "Objective Value": summary.objective_value
            if summary.objective_value is not None
            else "",
            "Result Perf": summary.result_performance
            if summary.result_performance is not None
            else "",
        }

        for key in dynamic_tiling_keys:
            data[key] = summary.tiling_values.get(key, "")

        for key in fixed_tiling_keys:
            data[key] = summary.tiling_values.get(key, "")

        for col_num, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_num, column=col_num, value=data[header])
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(fieldnames) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

    wb.save(output_file)
    print(f"Data exported to {output_file}")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Parse log files and generate summary tables"
    )
    parser.add_argument(
        "log_path", help="Path to log file or directory containing log files"
    )
    parser.add_argument(
        "-o", "--output", help="Output file path (optional, default: print to console)"
    )
    parser.add_argument(
        "-a",
        "--all",
        action="store_true",
        help="Summarize best group/case tiling data for all results",
    )
    parser.add_argument(
        "-f",
        "--format",
        choices=["console", "csv", "excel"],
        default="console",
        help="Output format: console (default), csv, or excel",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Recursively find all .log files in directory",
    )

    args = parser.parse_args()

    if not os.path.exists(args.log_path):
        print(f"Error: Path '{args.log_path}' not found.")
        return

    if os.path.isdir(args.log_path) or args.recursive:
        log_files = find_log_files(args.log_path)
        if not log_files:
            print(f"No .log files found in '{args.log_path}'")
            return
        print(f"Found {len(log_files)} log file(s)")
    else:
        log_files = [args.log_path]

    log_parser = LogParser()
    all_summaries = []

    for log_file in log_files:
        print(f"Parsing: {log_file}")
        try:
            summary_mode = (
                "all_results_all_groups" if args.all else "best_result_all_groups"
            )
            summaries = log_parser.parse_log_file(log_file, summary_mode=summary_mode)
            all_summaries.extend(summaries)
        except Exception as e:
            print(f"Warning: Failed to parse {log_file}: {e}")

    print(f"Total operators extracted: {len(all_summaries)}")

    if not all_summaries:
        print("No data to export.")
        return

    if args.format == "console":
        if args.output:
            with open(args.output, "w", encoding="utf-8") as f:
                original_stdout = sys.stdout
                sys.stdout = f
                print_summary_table(all_summaries)
                sys.stdout = original_stdout
            print(f"Summary table has been written to '{args.output}'")
        else:
            print_summary_table(all_summaries)
    elif args.format == "csv":
        output_file = args.output if args.output else "output.csv"
        export_to_csv(all_summaries, output_file)
        print(f"Data exported to {output_file}")
    elif args.format == "excel":
        output_file = args.output if args.output else "output.xlsx"
        export_to_excel(all_summaries, output_file)


if __name__ == "__main__":
    main()
